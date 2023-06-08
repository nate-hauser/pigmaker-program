import textractprettyprinter
import textractcaller
import pandas as pd
pd.options.mode.chained_assignment = None
from trp import Document
import numpy as np
from pdf2image import convert_from_path
from openpyxl import load_workbook
from textractcaller.t_call import call_textract, Textract_Features
from textractprettyprinter.t_pretty_print import convert_table_to_list
import pickle
from fpdf import FPDF
from dateutil.parser import ParserError
import pd_table as pdt
import boto3
from botocore.config import Config
from PIL import Image
import pytesseract
from img2table.document import PDF
from img2table.ocr import TesseractOCR
from datetime import timedelta
import tempfile

code_dict = {22: "Savaged", 21: "Ruptures", 27: "Starvation"}
for key in [1, 13, 20, 26]:
    code_dict[key] = "Low Viability"
for key in [10, 12, 28]:
    code_dict[key] = "Laid On"
for key in [23, 30]:
    code_dict[key] = "Strep"
for key in [2, 3, 7, 9, 15, 25, 31]:
    code_dict[key] = "Other"
# Scanning pngs using Textract


def extract_data(input_document):
    with tempfile.TemporaryDirectory() as temp_dir:
        pages = convert_from_path(input_document,output_folder=temp_dir,
                                  poppler_path=r"C:\Users\jakeh\poppler-0.68.0_x86\poppler-0.68.0\bin")
        dfs = list()
        client = boto3.client("textract", config=Config(connect_timeout=6000))
        for i, page in enumerate(pages):
            # if i != 0:
                page.save("out" + str(i) + ".png", "PNG")
                resp = call_textract(input_document="out" + str(i) + ".png", features=[Textract_Features.TABLES], boto3_textract_client=client)
                tdoc = Document(resp)
                dfs.append(pd.DataFrame(convert_table_to_list(trp_table=tdoc.pages[0].tables[0])))

    return pd.concat(dfs)

def extract_data2(input_document):
    pages = convert_from_path(input_document,
                              poppler_path=r"C:\Users\jakeh\poppler-0.68.0_x86\poppler-0.68.0\bin")
    dfs = pd.DataFrame()
    for i, page in enumerate(pages):
        page.save("out" + str(i) + ".png", "PNG")
        print(pytesseract.image_to_data(Image.open('out'+str(i)+".png"),output_type='data.frame').to_string())

    return dfs

def extract_data3(input_document):
    pdf=PDF(src=input_document)

    ocr = TesseractOCR(lang="eng")

    pdf_tables = pdf.extract_tables(ocr=ocr)
    for i in pdf_tables:
        print(i.to_string())


# Data Cleaning
def general_clean(df1):
    df1.reset_index(drop=True, inplace=True)
    df1 = df1.applymap(lambda x: x.strip() if type(x) == str else x)
    # df1.replace("NOT_SELECTED,",np.NaN,inplace=True)
    df1.rename(columns=df1.iloc[0], inplace=True)
    df1.rename(columns={'':"Crate#"},inplace=True)
    df1.drop(df1[df1['Sow ID'] == ''].index, inplace=True)
    df1.reset_index(drop=True, inplace=True)
    df1.columns = df1.columns.str.strip()
    df1.drop(df1[df1['Sow ID'] == 'Sow ID'].index, inplace=True)
    df1.replace('', np.NaN, inplace=True)
    df1.reset_index(drop=True, inplace=True)
    return df1


# Convert date columns to datetime data type
def convert_to_date(df1, column_list, start_end_dates):
    for x in column_list:
        if not x == "LW":
            possible_dates, possible_days, four_digit_dates = generate_possible_dates(x, start_end_dates)

            for i, value in enumerate(df1[x]):
                if pd.notna(value):
                    value=str(value)
                    if len(value)<=2:
                        converted_value_location = possible_days.index(int(value))
                        df1.at[i,x] = possible_dates[converted_value_location]
                    elif len(value) == 4:
                        converted_value_location = four_digit_dates.index(value)
                        df1.at[i,x] = possible_dates[converted_value_location]

        else:
            for i, value in enumerate(df1[x]):
                df1.at[i,x] = pd.to_datetime(value).date()

    return df1


def convert_to_numeric(df1, columns_list):

    for x in columns_list:
        for i, value in enumerate(df1[x]):
            df1.at[i, x] = pd.to_numeric(value)

    return df1

def produce_numeric_errors(df,cols_list):
    numeric_error_list=[]
    for x in cols_list:
        for i, value in enumerate(df[x]):
            try:
                pd.to_numeric(value)

            except ParserError as pe:
                numeric_error_list.append([i,df.columns.get_loc(x)])

            except ValueError as e:
                numeric_error_list.append([i,df.columns.get_loc(x)])

    return numeric_error_list

def generate_possible_dates(col,start_end_dates):

    if "Bred" in col:
        start = pd.to_datetime(start_end_dates["breed start"]).date()
        end = pd.to_datetime(start_end_dates["breed end"]).date()
    elif "Farrowed" in col:
        start = pd.to_datetime(start_end_dates["farrow start"]).date()
        end = pd.to_datetime(start_end_dates["farrow end"]).date()
    else:
        start = pd.to_datetime(start_end_dates["wean start"]).date()
        end = pd.to_datetime(start_end_dates["wean end"]).date()

    possible_dates = []
    num_days = (end - start).days

    for i in range(num_days + 1):
        possible_dates.append(start + timedelta(days=i))

    possible_days = [y.day for y in possible_dates]

    four_digit_dates = [str(y.month) + str(y.day) if len(str(y.day)) == 2 else str(y.month) + "0" + str(y.day) for y in
                        possible_dates]
    four_digit_dates = [y if len(y) == 4 else "0" + y for y in four_digit_dates]

    return possible_dates, possible_days, four_digit_dates

def produce_date_errors(df,cols_list, start_end_dates):
    dates_error_list = []
    for x in cols_list:
        if not x == "LW":
            possible_dates, possible_days, four_digit_dates = generate_possible_dates(x,start_end_dates)

            for i, value in enumerate(df[x]):
                try:
                    if pd.notna(value):
                        if len(value) <= 2:
                            if not int(value) in possible_days:
                                dates_error_list.append([i, df.columns.get_loc(x)])

                        elif len(value) == 4:
                            if not value in four_digit_dates:
                                dates_error_list.append([i, df.columns.get_loc(x)])

                        else:
                            dates_error_list.append([i,df.columns.get_loc(x)])

                except ParserError as pe:
                    dates_error_list.append([i, df.columns.get_loc(x)])
                except ValueError as e:
                    dates_error_list.append([i, df.columns.get_loc(x)])
        else:
            for i, value in enumerate(df[x]):
                try:
                    if pd.notna(value):
                        pd.to_datetime(value)

                except ValueError as e:
                    dates_error_list.append([i,df.columns.get_loc(x)])

    return dates_error_list


def fill_table(df1):

    for i in range(0, len(df1)):
        if pd.isna(df1.at[i, "Date Weaned"]) and pd.notna(df1.at[i, "Date Farrowed"]):
            df1.at[i, "Date Weaned"] = df1.at[i - 1, "Date Weaned"]

        for count in range (1,4):
            for i in range (0,len(df1)):
                if pd.isna(df1.at[i,"Breeder"+str(count)]) and pd.notna(df1.at[i,"HC"+str(count)]):
                    df1.at[i,"Breeder"+str(count)]=df1.at[i,"HC"+str(count)]
    # Waiting to get more info

    return df1

def breed_produce_errors(df1,start_end_dates):

    df1.replace({"NOT_SELECTED,":np.NaN},inplace=True)

    dates_cols_list = ["Date Bred1","Date Bred2","Date Bred3","LW"]
    breeder_hc_cols = ["HC1","Breeder1","HC2","Breeder2","HC3","Breeder3"]
    breeder_list = ["BV","AC","CJ","HR","JS","J","NS"]
    df1[breeder_hc_cols] = df1[breeder_hc_cols].replace({'AL': 'AC', 'BL': 'BV', 'PV': 'BV',
                                                          'B': 'BV', 'A': 'AC', 'H': 'HR',
                                                          "C": "CJ","J":"JS","P":"BV","8":"BV",
                                                         "it":"HR","lt":"HR","3":"BV","It":"HR","f":"BV",
                                                         "b":"BV"})
    df1[dates_cols_list] = df1[dates_cols_list].replace({"1b":"16","1>":"17","lb":"16","1)":"17","lt":"16","It":"16"})

    error_list = []

    for x in breeder_hc_cols:
        for i, value in enumerate(df1[x]):
            if value not in breeder_list and pd.notna(value):
                error_list.append([i,df1.columns.get_loc(x)])

    error_list.extend(produce_date_errors(df1,dates_cols_list,start_end_dates))
    error_list.extend(produce_sow_id_errors(df1))

    df1.to_pickle("breed2.pkl")


    return error_list


def farrow_produce_errors(df1,start_end_dates):
    global code_dict
    dates_cols_list = ["Date Farrowed","Date Weaned"]
    numeric_cols_list = ["Crate#","P","#L","#S","#M","#W"]
    error_dict = {r"\\": "1", "I": "1", "o": "0", "O": "0", "&": "9", "a": "9", "/": "1"}
    df1[numeric_cols_list] = df1[numeric_cols_list].replace(error_dict, regex=True)
    error_list = []
    for i in range(1, 5):
        for count, x in enumerate(df1["C" + str(i)]):
            is_num_deaths = True
            num_deaths = []
            death_code = []
            if pd.notna(x):
                for j in x:
                    if j == '-':
                        is_num_deaths = False

                    elif is_num_deaths:
                        num_deaths.append(j)

                    else:
                        death_code.append(j)

                deaths = ''.join(num_deaths)
                code = ''.join(death_code)
                try:
                    if int(code) not in code_dict.keys() or not deaths.isdigit():
                        error_list.append([count, df1.columns.get_loc("C" + str(i))])
                except ValueError:
                    error_list.append([count,df1.columns.get_loc("C"+str(i))])

    error_list.extend(produce_date_errors(df1,dates_cols_list,start_end_dates))
    error_list.extend(produce_numeric_errors(df1,numeric_cols_list))
    error_list.extend(produce_sow_id_errors(df1))

    df1.to_pickle("farrow2.pkl")

    return error_list

def produce_sow_id_errors(df):
    error_list = []
    for i, value in enumerate(df["Sow ID"]):
        if not value.isalnum() or not value.isupper() and not value.isdigit():
            error_list.append([i,df.columns.get_loc("Sow ID")])

    return error_list

def pdf_to_breed(filepath):
    df = extract_data(filepath)
    df = general_clean(df)
    # df.replace({'AL': 'AC', 'BL': 'BV', 'PV': 'BV', 'B': 'BV', 'A': 'AC', 'H': 'HR', "C": "CJ","J":"JS",
    #             "1b":"16","1>":"17","P":"BV","3":"BV","8":"BV","lb":"16"}, inplace=True)
    df.to_pickle("breed.pkl")
    return df

def pdf_to_farrow(filepath):
    df = extract_data(filepath)
    df = general_clean(df)

    return df

def pre_report_processing(df1,df2,start_end_dates):
    for i in range (0,len(df1)):
        if i == 0 and pd.isna(df1.at[i,"LW"]):
            break
        elif pd.isna(df1.at[i,"Last Weaned"]) and pd.notna(df1.at[i,"Date Bred1"]):
            df1.at[i,"Last Weaned"]=df1.at[i-1,"Last Weaned"]

    df1 = convert_to_date(df1,["Date Bred1","Date Bred2","Date Bred3","LW"],start_end_dates)
    df2 = convert_to_date(df2,["Date Farrowed","Date Weaned"],start_end_dates)
    df2 = convert_to_numeric(df2,["Crate#","P","#L","#S","#M","#W"])


    df2["Low Viability"] = 0
    df2["Laid On"] = 0
    df2["Strep"] = 0
    df2["Scours"] = 0
    df2["Other"] = 0
    df2["Savaged"] = 0
    df2["Ruptures"] = 0
    df2["Starvation"] = 0
    df2["Unknown"] = 0
    for i in range(1, 5):
        for count, x in enumerate(df2["C" + str(i)]):
            is_num_deaths = True
            num_deaths = []
            death_code = []
            if pd.notna(x):
                for j in x:
                    if j == '-':
                        is_num_deaths = False

                    elif is_num_deaths:
                        num_deaths.append(j)

                    else:
                        death_code.append(j)

                deaths = ''.join(num_deaths)
                code = ''.join(death_code)
                df2.at[count, code_dict[int(code)]] = int(deaths)
        df2.drop("C" + str(i), axis=1, inplace=True)

    df2[["#L","#S","#M","#W"]]=df2[["#L","#S","#M","#W"]].replace(np.NaN,0)

    df3 = df2.merge(df1, how='outer', on="Sow ID")

    return df3
def generate_report(df3,group_num):

    df3 = fill_table(df3)
    length=len(df3)
    df3=df3.replace("Ll",11)
    df3=df3.replace("J","JS")
    df3=df3.replace("BY","BV")
    df3=df3.replace("nan",np.NaN)
    # df3.insert(loc=17,column="Unknown",value=np.NaN)
    # for i, value in enumerate (df3["Date Farrowed"]):
    #     if pd.notna(value):
    #         df3.at[i,"Unknown"]=0

    df3["Group Number"] = int(group_num)
    diff = df3["#L"].sum() - (df3["Low Viability"].sum() + df3["Laid On"].sum() + df3["Strep"].sum() + df3["Scours"].sum()
                       + df3["Other"].sum() + df3["Savaged"].sum() + df3["Ruptures"].sum() + df3["Starvation"].sum())
    balance = 0
    if diff > df3["#W"].sum():
        df3 = pd.concat([df3,pd.DataFrame([{"Sow ID":"Y","Unknown":diff-df3["#W"].sum(),"Group Number":group_num}])],ignore_index=True)
        balance = 1
    elif diff < df3["#W"].sum():
        df3 = pd.concat([df3,pd.DataFrame([{"Sow ID": "Y", "#W": diff-df3["#W"].sum(),"Group Number":group_num}])], ignore_index=True)
        balance = 2

    # print(df3["#L"].sum()-df3["#W"].sum())
    # print(df3["Low Viability"].sum() + df3["Laid On"].sum() + df3["Strep"].sum() + df3["Scours"].sum()
    #                    + df3["Other"].sum() + df3["Savaged"].sum() + df3["Ruptures"].sum() + df3["Starvation"].sum() +df3["Unknown"].sum())
    pdf = FPDF()
    pdf.add_page()

    pdf.set_font("Times","B",25)
    pdf.cell(0,0,"Group "+group_num+ " Pigmaker Report",0,2,"C")

    mode=df3["Date Weaned"].mode()[0].strftime("%m-%d-%Y")
    pdf.set_y(27)
    pdf.cell(0,0,str(mode),0,2,'C')
    pdf.set_font("Times",'B',15)

    pdf.set_y(40)

    pdf.cell(48,8,"Statistic",1,0,'C')
    pdf.cell(48,8,"Rooms 1-6",1,0,'C')
    pdf.cell(48,8,"Rooms 8-12",1,0,'C')
    pdf.cell(48,8,"Overall",1,1,'C')

    pdf.set_font("Times", '', 12)

    pdf.cell(48,8,"Total born per litter",1,1,'C')
    pdf.cell(48, 8,"Still birth rate", 1, 1,'C')
    pdf.cell(48, 8, "Mum birth rate", 1, 1, 'C')
    pdf.cell(48, 8, "Born live per litter", 1, 1, 'C')
    pdf.cell(48, 8, "Weaned per litter", 1, 1, 'C')
    pdf.cell(48, 8, "Pre-weaning mortality rate", 1, 1, 'C')

    pdf.set_xy(58,48)

    num_farrowed = df3["Date Farrowed"].count()
    farrowing_rate = num_farrowed / length

    Half1 = df3[df3["Crate#"] < 700]
    Half2 = df3[df3["Crate#"] >= 800]
    for half in [Half1,Half2]:
        half_num_farrowed = half["Date Farrowed"].count()

        total_born = round(((half["#S"].sum() + half["#M"].sum() + half["#L"].sum())/half_num_farrowed),2)
        pdf.cell(48,8,str(total_born),1,2,'C')

        still_rate=half["#S"].sum()/(total_born * half_num_farrowed)
        pdf.cell(48,8,"{:.2%}".format(still_rate),1,2,'C')

        mum_rate=half["#M"].sum()/(total_born * half_num_farrowed)
        pdf.cell(48,8,"{:.2%}".format(mum_rate),1,2,'C')

        live_per_litter=round(half["#L"].sum()/half_num_farrowed,2)
        pdf.cell(48, 8, str(live_per_litter), 1, 2,'C')

        weaned_per_litter=round(half["#W"].sum()/half_num_farrowed,2)
        pdf.cell(48, 8, str(weaned_per_litter), 1, 2,'C')

        preweaning_mortality = (half["Low Viability"].sum() + half["Laid On"].sum() + half["Strep"].sum() + half["Scours"].sum()
        + half["Other"].sum() + half["Savaged"].sum() + half["Ruptures"].sum() + half["Starvation"].sum())/half["#L"].sum()
        pdf.cell(48, 8, "{:.2%}".format(preweaning_mortality), 1, 2,'C')

        pdf.set_xy(106,48)

    pdf.set_x(154)

    total_num_farrowed = df3["Date Farrowed"].count()

    total_born = round(((df3["#S"].sum() + df3["#M"].sum() + df3["#L"].sum()) / total_num_farrowed), 2)
    pdf.cell(48, 8, str(total_born), 1, 2, 'C')

    still_rate = df3["#S"].sum() / (total_born * total_num_farrowed)
    pdf.cell(48, 8, "{:.2%}".format(still_rate), 1, 2, 'C')

    mum_rate = df3["#M"].sum() / (total_born * total_num_farrowed)
    pdf.cell(48, 8, "{:.2%}".format(mum_rate), 1, 2, 'C')

    live_per_litter = round(df3["#L"].sum() / total_num_farrowed, 2)
    pdf.cell(48, 8, str(live_per_litter), 1, 2, 'C')

    weaned_per_litter = round(df3["#W"].sum() / total_num_farrowed, 2)
    pdf.cell(48, 8, str(weaned_per_litter), 1, 2, 'C')

    preweaning_mortality = (df3["Low Viability"].sum() + df3["Laid On"].sum() + df3["Strep"].sum() + df3[
        "Scours"].sum() + df3["Other"].sum() + df3["Savaged"].sum() + df3["Ruptures"].sum() + df3[
                                "Starvation"].sum()+df3["Unknown"].sum()) / df3["#L"].sum()
    pdf.cell(48, 8, "{:.2%}".format(preweaning_mortality), 1, 2, 'C')


    pdf.set_xy(10,105)

    pdf.cell(0,10,"*The overall farrowing rate is: "+"{:.2%}".format(farrowing_rate)+" ("+str(num_farrowed)+" out of "+str(length)+")",0,1,'L')
    weaned_pigs=df3["#W"].sum()
    pdf.cell(0,10,"*The total number of weaned pigs is "+str(weaned_pigs),0,1,'L')
    if balance == 1:
        pdf.multi_cell(0,10,"*"+str(int(diff-df3["#W"].sum()))+
        " unknown deaths were added to imaginary sow 'Y' to make the total born live minus the total dead equal the total weaned",0,'L')
    elif balance ==2:
        pdf.multi_cell(0,10, "*" + str(int(df3["#W"].sum() - diff)) +
        " weaned pigs were added to imaginary sow 'Y' to make the total born live minus the total dead equal the total weaned", 0, 'L')
    else:
        pdf.multi_cell(0,10, "* The total born live minus the total dead equal the total weaned", 0, 'L')



    combo=df3.copy()
    combo["HC_combo"] = combo.apply(lambda row: list({row['HC1'], row['HC2'], row['HC3']}), axis=1)
    combo['HC_combo'] = combo['HC_combo'].apply(lambda x: [i for i in x if not (pd.isna(i) or i is None)]).astype('object')
    combo["Breeder_combo"]= combo.apply(lambda row: list({row['Breeder1'], row['Breeder2'], row['Breeder3']}), axis=1)
    combo['Breeder_combo'] = combo['Breeder_combo'].apply(lambda x: [i for i in x if not (pd.isna(i) or i is None)]).astype('object')

    combo['HC_combo2'] = combo['HC_combo'].apply(tuple)
    combo['Breeder_combo2'] = combo['Breeder_combo'].apply(tuple)

    hc = combo.groupby(by="HC_combo2").agg({'#S':'sum',
                                           '#M':'sum',
                                           '#L':'sum',
                                           'Group Number':'count',
                                           'Date Farrowed':'count'})

    hc["Farrowing rate"] = hc["Date Farrowed"]/hc["Group Number"]
    hc["Total born"] = hc["#S"] + hc["#M"] + hc["#L"]
    hc = hc.sort_values(by="Farrowing rate",ascending=False)


    pdf.set_font('Times','B',12)

    pdf.set_y(150)
    pdf.cell(49,8,"Heat checker combination",1,0,'C')
    pdf.cell(49, 8, "Total born per litter", 1, 0, 'C')
    pdf.cell(49, 8, "Farrowing rate", 1, 0, 'C')
    pdf.cell(49, 8, "Count", 1, 1, 'C')

    pdf.set_font('Times','',12)

    for i in hc.index:
        if i == ():
            pdf.cell(49,8,"Unknown",1,0,'C')
        else:
            pdf.cell(49,8,str(', '.join(sorted(i))),1,0,'C')

        if hc.at[i,"Date Farrowed"] == 0:
            pdf.cell(49,8,"NA",1,0,'C')
        else:
            pdf.cell(49, 8, str(round(hc.at[i,"Total born"]/hc.at[i,"Date Farrowed"],2)), 1, 0, 'C')
        pdf.cell(49, 8, "{:.2%}".format(hc.at[i,"Farrowing rate"]), 1, 0, 'C')
        pdf.cell(49,8,str(hc.at[i,"Group Number"]),1,1,'C')


    pdf.set_y(pdf.get_y()+10)

    pdf.set_font('Times','B',12)
    pdf.cell(49,8,"Individual heat checker",1,0,'C')
    pdf.cell(49, 8, "Total born per litter", 1, 0, 'C')
    pdf.cell(49, 8, "Farrowing rate", 1, 0, 'C')
    pdf.cell(49, 8, "Count", 1, 1, 'C')

    pdf.set_font('Times', '', 12)

    v = []
    for i in hc.index:
        v.extend(i)

    v = list(set(v))
    dfl = {}
    for i in v:
        i_df = hc[hc.index.map(lambda tpl: any(str(i) in elem for elem in tpl))]
        dfl[str(i)]=(i_df["Date Farrowed"].sum() / i_df["Group Number"].sum())

    sorted_dfl = sorted(dfl.items(), key = lambda kv: kv[1],reverse = True)
    heat_checkers = []
    for heat_checker, farrowing_rate in sorted_dfl:
        heat_checkers.append(heat_checker)
    for i in heat_checkers:
        i_df = hc[hc.index.map(lambda tpl: any(str(i) in elem for elem in tpl))]
        pdf.cell(49,8,str(i),1,0,'C')
        if i_df["Date Farrowed"].sum() == 0:
            pdf.cell(49,8,"NA",1,0,'C')
        else:
            tbpl = round(i_df["Total born"].sum()/i_df["Date Farrowed"].sum(),2)
            pdf.cell(49,8,str(tbpl),1,0,'C')

        pdf.cell(49, 8, "{:.2%}".format(i_df["Date Farrowed"].sum()/i_df["Group Number"].sum()), 1, 0, 'C')
        pdf.cell(49,8,str(i_df["Group Number"].sum()),1,1,'C')

    br = combo.groupby(by="Breeder_combo2").agg({'#S': 'sum',
                                                 '#M': 'sum',
                                                 '#L': 'sum',
                                                 'Group Number': 'count',
                                                 'Date Farrowed': 'count'})

    br["Farrowing rate"] = br["Date Farrowed"] / br["Group Number"]
    br["Total born"] = br["#S"] + br["#M"] + br["#L"]

    br = br.sort_values(by="Farrowing rate", ascending=False)

    pdf.set_y(pdf.get_y()+10)

    pdf.set_font('Times','B',12)
    pdf.cell(49,8,"Breeder combination",1,0,'C')
    pdf.cell(49, 8, "Total born per litter", 1, 0, 'C')
    pdf.cell(49, 8, "Farrowing rate", 1, 0, 'C')
    pdf.cell(49, 8, "Count", 1, 1, 'C')

    pdf.set_font('Times','',12)

    for i in br.index:
        if i == ():
            pdf.cell(49, 8, "Unknown", 1, 0, 'C')
        else:
            pdf.cell(49, 8, str(', '.join(sorted(i))), 1, 0, 'C')

        if br.at[i, "Date Farrowed"] == 0:
            pdf.cell(49, 8, "NA", 1, 0, 'C')
        else:
            pdf.cell(49, 8, str(round(br.at[i, "Total born"] / br.at[i, "Date Farrowed"], 2)), 1, 0, 'C')
        pdf.cell(49, 8, "{:.2%}".format(br.at[i, "Farrowing rate"]), 1, 0, 'C')
        pdf.cell(49, 8, str(br.at[i, "Group Number"]), 1, 1, 'C')

    pdf.set_y(pdf.get_y() + 10)

    pdf.set_font('Times', 'B', 12)
    pdf.cell(49, 8, "Individual breeder", 1, 0, 'C')
    pdf.cell(49, 8, "Total born per litter", 1, 0, 'C')
    pdf.cell(49, 8, "Farrowing rate", 1, 0, 'C')
    pdf.cell(49, 8, "Count", 1, 1, 'C')

    pdf.set_font('Times', '', 12)

    v = []
    for i in br.index:
        v.extend(i)

    v = list(set(v))
    dfl = {}
    for i in v:
        i_df = br[br.index.map(lambda tpl: any(str(i) in elem for elem in tpl))]
        dfl[str(i)] = (i_df["Date Farrowed"].sum() / i_df["Group Number"].sum())

    sorted_dfl = sorted(dfl.items(), key=lambda kv: kv[1], reverse=True)
    breeders = []
    for breeder, farrowing_rate in sorted_dfl:
        breeders.append(breeder)
    for i in breeders:

        i_df = br[br.index.map(lambda tpl: any(str(i) in elem for elem in tpl))]
        pdf.cell(49, 8, str(i), 1, 0, 'C')
        if i_df["Date Farrowed"].sum() == 0:
            pdf.cell(49, 8, "NA", 1, 0, 'C')
        else:
            tbpl = round(i_df["Total born"].sum() / i_df["Date Farrowed"].sum(), 2)
            pdf.cell(49, 8, str(tbpl), 1, 0, 'C')

        pdf.cell(49, 8, "{:.2%}".format(i_df["Date Farrowed"].sum() / i_df["Group Number"].sum()), 1, 0, 'C')
        pdf.cell(49, 8, str(i_df["Group Number"].sum()), 1, 1, 'C')






    pdf.output("Group" + group_num + " Report.pdf")
    print("Report has been created")

    return df3

def output_to_excel(df):
    reader = pd.read_excel(r'C:\Users\jakeh\output.xlsx')
    writer = pd.ExcelWriter(r"C:\Users\jakeh\output.xlsx", engine='openpyxl', mode='a', if_sheet_exists="overlay")
    df.to_excel(writer,header=False,index=False,startrow=len(reader)+1)
    writer.close()
